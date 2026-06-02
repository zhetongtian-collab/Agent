from html import escape
from math import isfinite
from pathlib import Path
import re
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from app.core.config import settings


# 分析 Excel 文件的结构，返回每个工作表的概览信息。
# max_rows 控制每个工作表最多展示多少行样例数据。
# 返回结果包括工作表名、非空行数、列数、表头和样例行，
# 方便 Agent 在不读取完整大表的情况下先了解表格长什么样。
def analyze_excel_file(path: str | Path, max_rows: int = 5) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            non_empty_rows = [
                ["" if value is None else value for value in row]
                for row in rows
                if any(value is not None and str(value).strip() for value in row)
            ]
            headers = [str(value) for value in non_empty_rows[0]] if non_empty_rows else []
            samples = [
                ["" if value is None else str(value) for value in row]
                for row in non_empty_rows[1 : max_rows + 1]
            ]
            sheets.append(
                {
                    "sheet_name": sheet.title,
                    "row_count": len(non_empty_rows),
                    "column_count": max((len(row) for row in non_empty_rows), default=0),
                    "headers": headers,
                    "sample_rows": samples,
                }
            )
        return {"sheets": sheets}
    finally:
        workbook.close()


def read_excel_range_data(
    path: str | Path,
    cell_range: str,
    sheet_name: str | None = None,
    data_only: bool = True,
    max_cells: int = 2000,
) -> dict[str, Any]:
    normalized_range = _validate_cell_range(cell_range)
    min_column, min_row, max_column, max_row = range_boundaries(normalized_range)
    cell_count = (max_column - min_column + 1) * (max_row - min_row + 1)
    if cell_count > max_cells:
        raise ValueError(f"区域过大：{cell_count} 个单元格，最多读取 {max_cells} 个")
    workbook = load_workbook(path, data_only=data_only, read_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = [
            [sheet.cell(row=row, column=column).value for column in range(min_column, max_column + 1)]
            for row in range(min_row, max_row + 1)
        ]
        return {"sheet_name": sheet.title, "range": normalized_range, "rows": rows}
    finally:
        workbook.close()


def sum_excel_range(
    path: str | Path,
    sum_range: str,
    sheet_name: str | None = None,
    criteria_range: str | None = None,
    criteria: Any | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        sum_values = _range_values(sheet, sum_range)
        if criteria_range:
            criteria_values = _range_values(sheet, criteria_range)
            if len(criteria_values) != len(sum_values):
                raise ValueError("criteria_range 和 sum_range 必须包含相同数量的单元格")
            selected = [
                value
                for value, candidate in zip(sum_values, criteria_values)
                if _matches_filter(candidate, "eq", criteria)
            ]
        else:
            selected = sum_values
        total = sum(number for value in selected if (number := _parse_number(value)) is not None)
        return {"sheet_name": sheet.title, "sum_range": sum_range, "criteria": criteria, "result": total}
    finally:
        workbook.close()


def lookup_excel_value(
    path: str | Path,
    lookup_value: Any,
    lookup_range: str,
    result_range: str,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        lookup_values = _range_values(sheet, lookup_range)
        result_values = _range_values(sheet, result_range)
        if len(lookup_values) != len(result_values):
            raise ValueError("lookup_range 和 result_range 必须包含相同数量的单元格")
        for index, candidate in enumerate(lookup_values):
            if _matches_filter(candidate, "eq", lookup_value):
                return {
                    "sheet_name": sheet.title,
                    "lookup_value": lookup_value,
                    "matched_index": index,
                    "result": result_values[index],
                }
        return {"sheet_name": sheet.title, "lookup_value": lookup_value, "matched_index": None, "result": None}
    finally:
        workbook.close()


def filter_excel_rows(
    path: str | Path,
    data_range: str,
    column: str,
    operator: str,
    value: Any,
    sheet_name: str | None = None,
    limit: int = 200,
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = _range_rows(sheet, data_range)
        if not rows:
            return {"sheet_name": sheet.title, "range": data_range, "rows": []}
        headers = ["" if item is None else str(item) for item in rows[0]]
        column_index = _resolve_column(column, headers, default=0)
        matches = [
            row
            for row in rows[1:]
            if column_index < len(row) and _matches_filter(row[column_index], operator, value)
        ]
        return {
            "sheet_name": sheet.title,
            "range": data_range,
            "headers": headers,
            "matched_count": len(matches),
            "rows": matches[:limit],
            "truncated": len(matches) > limit,
        }
    finally:
        workbook.close()


def generate_excel_chart_image(
    path: str | Path,
    chart_type: str = "line",
    sheet_name: str | None = None,
    x_axis_column: str | None = None,
    y_columns: list[str] | None = None,
    title: str | None = None,
    max_rows: int = 50,
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        table = _read_table(sheet, max_rows=max_rows)
    finally:
        workbook.close()

    if len(table) < 2:
        raise ValueError("Excel 工作表没有足够的数据行，无法生成图表")

    headers = _normalize_headers(table[0])
    data_rows = table[1:]
    x_index = _resolve_column(x_axis_column, headers, default=0)
    y_indexes = _resolve_y_columns(y_columns, headers, data_rows, x_index)
    if not y_indexes:
        raise ValueError("未找到可用于生成图表的数值列")

    labels: list[str] = []
    series = [{"name": headers[index], "values": []} for index in y_indexes]
    for row in data_rows:
        labels.append(_cell_to_label(row[x_index] if x_index < len(row) else ""))
        for item, column_index in zip(series, y_indexes):
            value = row[column_index] if column_index < len(row) else None
            item["values"].append(_parse_number(value))

    if not any(value is not None for item in series for value in item["values"]):
        raise ValueError("选中的列不包含可用数值")

    normalized_chart_type = chart_type.lower().strip()
    if normalized_chart_type not in {"line", "bar"}:
        raise ValueError("chart_type 必须是 line 或 bar")

    settings.artifact_dir.mkdir(parents=True, exist_ok=True)
    source_stem = _safe_stem(Path(path).stem)
    image_path = settings.artifact_dir / f"{source_stem}-{normalized_chart_type}-chart-{uuid4().hex[:8]}.svg"
    chart_title = title or f"{sheet.title}{_chart_type_label(normalized_chart_type)}"
    image_path.write_text(
        _render_chart_svg(chart_title, labels, series, normalized_chart_type),
        encoding="utf-8",
    )

    return {
        "path": str(image_path),
        "chart": {
            "chart_type": normalized_chart_type,
            "sheet_name": sheet.title,
            "x_axis_column": headers[x_index],
            "y_columns": [headers[index] for index in y_indexes],
            "row_count": len(data_rows),
            "title": chart_title,
        },
    }


def _read_table(sheet: Any, max_rows: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        trimmed = _trim_row(list(row))
        if any(value is not None and str(value).strip() for value in trimmed):
            rows.append(trimmed)
        if len(rows) >= max_rows + 1:
            break
    return rows


def _range_rows(sheet: Any, cell_range: str) -> list[list[Any]]:
    normalized_range = _validate_cell_range(cell_range)
    min_column, min_row, max_column, max_row = range_boundaries(normalized_range)
    return [
        [sheet.cell(row=row, column=column).value for column in range(min_column, max_column + 1)]
        for row in range(min_row, max_row + 1)
    ]


def _range_values(sheet: Any, cell_range: str) -> list[Any]:
    return [value for row in _range_rows(sheet, cell_range) for value in row]


def _validate_cell_range(cell_range: str) -> str:
    normalized = str(cell_range or "").strip().upper()
    if "!" in normalized:
        _, normalized = normalized.rsplit("!", 1)
    try:
        min_column, min_row, max_column, max_row = range_boundaries(normalized)
    except (IndexError, TypeError, ValueError) as exc:
        raise ValueError("Excel 区域格式无效，请使用完整区域，例如 A1:F20") from exc
    if None in {min_column, min_row, max_column, max_row}:
        raise ValueError("Excel 区域必须同时包含起止行和列，例如 A1:F20")
    return normalized


def _matches_filter(candidate: Any, operator: str, expected: Any) -> bool:
    normalized_operator = operator.strip().lower()
    if normalized_operator in {"eq", "="}:
        return _normalize_compare(candidate) == _normalize_compare(expected)
    if normalized_operator in {"neq", "!=", "<>"}:
        return _normalize_compare(candidate) != _normalize_compare(expected)
    if normalized_operator == "contains":
        return _normalize_compare(expected) in _normalize_compare(candidate)
    candidate_number = _parse_number(candidate)
    expected_number = _parse_number(expected)
    if candidate_number is None or expected_number is None:
        return False
    if normalized_operator in {"gt", ">"}:
        return candidate_number > expected_number
    if normalized_operator in {"gte", ">="}:
        return candidate_number >= expected_number
    if normalized_operator in {"lt", "<"}:
        return candidate_number < expected_number
    if normalized_operator in {"lte", "<="}:
        return candidate_number <= expected_number
    raise ValueError(f"不支持的筛选操作符：{operator}")


def _normalize_compare(value: Any) -> str:
    return "" if value is None else re.sub(r"\s+", "", str(value).strip().casefold())


def _trim_row(row: list[Any]) -> list[Any]:
    while row and (row[-1] is None or str(row[-1]).strip() == ""):
        row.pop()
    return row


def _normalize_headers(row: list[Any]) -> list[str]:
    headers: list[str] = []
    for index, value in enumerate(row, start=1):
        label = _cell_to_label(value).strip()
        headers.append(label or f"第 {index} 列")
    return headers


def _resolve_column(value: str | None, headers: list[str], default: int) -> int:
    if value is None or not str(value).strip():
        return default
    text = str(value).strip()
    if text.isdigit():
        index = int(text) - 1
        if 0 <= index < len(headers):
            return index
    column_letter = _column_letter_to_index(text)
    if column_letter is not None and 0 <= column_letter - 1 < len(headers):
        return column_letter - 1
    normalized = _normalize_header(text)
    for index, header in enumerate(headers):
        if _normalize_header(header) == normalized:
            return index
    raise ValueError(f"未找到列：{value}")


def _resolve_y_columns(
    y_columns: list[str] | None,
    headers: list[str],
    data_rows: list[list[Any]],
    x_index: int,
) -> list[int]:
    if y_columns:
        indexes = [_resolve_column(column, headers, default=0) for column in y_columns]
        return [index for index in indexes if index != x_index]

    numeric_indexes: list[int] = []
    for index in range(len(headers)):
        if index == x_index:
            continue
        if any(_parse_number(row[index] if index < len(row) else None) is not None for row in data_rows):
            numeric_indexes.append(index)
    return numeric_indexes[:4]


def _render_chart_svg(
    title: str,
    labels: list[str],
    series: list[dict[str, Any]],
    chart_type: str,
) -> str:
    width = 920
    height = 520
    left = 72
    right = 36
    top = 68
    bottom = 104
    plot_width = width - left - right
    plot_height = height - top - bottom
    plot_bottom = top + plot_height
    colors = ["#256f56", "#2f6fb3", "#c26a2c", "#9f3d46", "#6955a3"]
    numeric_values = [
        value
        for item in series
        for value in item["values"]
        if value is not None and isfinite(value)
    ]
    minimum = min(0, min(numeric_values))
    maximum = max(0, max(numeric_values))
    if minimum == maximum:
        minimum -= 1
        maximum += 1

    def x_pos(index: int) -> float:
        if len(labels) <= 1:
            return left + plot_width / 2
        return left + (plot_width * index / (len(labels) - 1))

    def y_pos(value: float) -> float:
        return plot_bottom - ((value - minimum) / (maximum - minimum) * plot_height)

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img">',
        "<style>",
        "text{font-family:Inter,'Segoe UI','Microsoft YaHei',Arial,sans-serif;fill:#243129} .axis{stroke:#53605a;stroke-width:1.2} .grid{stroke:#d8ded5;stroke-width:1} .label{font-size:12px;fill:#657169} .title{font-size:22px;font-weight:700;fill:#17211b} .legend{font-size:13px;font-weight:600}",
        "</style>",
        '<rect width="100%" height="100%" rx="8" fill="#fbfcf8"/>',
        f'<text class="title" x="{left}" y="36">{escape(title)}</text>',
    ]

    for tick in range(6):
        ratio = tick / 5
        value = minimum + (maximum - minimum) * ratio
        y = plot_bottom - plot_height * ratio
        parts.append(f'<line class="grid" x1="{left}" y1="{y:.2f}" x2="{left + plot_width}" y2="{y:.2f}"/>')
        parts.append(f'<text class="label" x="{left - 10}" y="{y + 4:.2f}" text-anchor="end">{_format_number(value)}</text>')

    parts.append(f'<line class="axis" x1="{left}" y1="{top}" x2="{left}" y2="{plot_bottom}"/>')
    parts.append(f'<line class="axis" x1="{left}" y1="{plot_bottom}" x2="{left + plot_width}" y2="{plot_bottom}"/>')

    if chart_type == "bar":
        parts.extend(_render_bar_series(labels, series, colors, left, plot_width, plot_bottom, y_pos))
    else:
        parts.extend(_render_line_series(labels, series, colors, x_pos, y_pos))

    label_step = max(1, len(labels) // 12)
    for index, label in enumerate(labels):
        if index % label_step != 0 and index != len(labels) - 1:
            continue
        x = x_pos(index)
        parts.append(
            f'<text class="label" x="{x:.2f}" y="{plot_bottom + 22}" text-anchor="end" transform="rotate(-35 {x:.2f} {plot_bottom + 22})">{escape(_shorten(label, 18))}</text>'
        )

    legend_x = left
    legend_y = height - 24
    for index, item in enumerate(series):
        x = legend_x + index * 170
        color = colors[index % len(colors)]
        parts.append(f'<rect x="{x}" y="{legend_y - 10}" width="12" height="12" rx="2" fill="{color}"/>')
        parts.append(f'<text class="legend" x="{x + 18}" y="{legend_y}">{escape(_shorten(str(item["name"]), 18))}</text>')

    parts.append("</svg>")
    return "\n".join(parts)


def _render_line_series(
    labels: list[str],
    series: list[dict[str, Any]],
    colors: list[str],
    x_pos: Any,
    y_pos: Any,
) -> list[str]:
    parts: list[str] = []
    for series_index, item in enumerate(series):
        color = colors[series_index % len(colors)]
        points = []
        for index, value in enumerate(item["values"]):
            if value is not None:
                points.append(f"{x_pos(index):.2f},{y_pos(value):.2f}")
        if len(points) >= 2:
            parts.append(
                f'<polyline fill="none" stroke="{color}" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" points="{" ".join(points)}"/>'
            )
        for index, value in enumerate(item["values"]):
            if value is not None:
                parts.append(f'<circle cx="{x_pos(index):.2f}" cy="{y_pos(value):.2f}" r="4" fill="{color}"/>')
    return parts


def _render_bar_series(
    labels: list[str],
    series: list[dict[str, Any]],
    colors: list[str],
    left: int,
    plot_width: int,
    plot_bottom: int,
    y_pos: Any,
) -> list[str]:
    parts: list[str] = []
    group_width = plot_width / max(1, len(labels))
    bar_width = min(44, group_width / (len(series) + 0.7))
    zero_y = y_pos(0)
    for label_index, _label in enumerate(labels):
        group_left = left + group_width * label_index + (group_width - bar_width * len(series)) / 2
        for series_index, item in enumerate(series):
            value = item["values"][label_index]
            if value is None:
                continue
            y = y_pos(value)
            rect_y = min(y, zero_y)
            rect_height = abs(zero_y - y)
            x = group_left + series_index * bar_width
            color = colors[series_index % len(colors)]
            parts.append(
                f'<rect x="{x:.2f}" y="{rect_y:.2f}" width="{max(2, bar_width - 3):.2f}" height="{max(1, rect_height):.2f}" rx="3" fill="{color}"/>'
            )
    return parts


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _cell_to_label(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _column_letter_to_index(value: str) -> int | None:
    letters = value.strip().upper()
    if not re.fullmatch(r"[A-Z]+", letters):
        return None
    index = 0
    for letter in letters:
        index = index * 26 + (ord(letter) - ord("A") + 1)
    return index


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", value.strip().lower())


def _safe_stem(filename: str) -> str:
    cleaned = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", filename).strip("_")
    return cleaned[:64] or "图表"


def _format_number(value: float) -> str:
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    if abs(value) >= 10:
        return f"{value:.1f}".rstrip("0").rstrip(".")
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _shorten(value: str, limit: int) -> str:
    return value if len(value) <= limit else f"{value[: limit - 3]}..."


def _chart_type_label(chart_type: str) -> str:
    return "折线图" if chart_type == "line" else "柱状图"
