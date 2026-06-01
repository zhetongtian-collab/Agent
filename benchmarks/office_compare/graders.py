from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any
from urllib import request

from docx import Document
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from benchmarks.office_compare.adapters import ChatResult


@dataclass
class Grade:
    passed: bool
    detail: str


def _normalized(value: Any) -> str:
    text = str(value).casefold()
    text = re.sub(
        r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?",
        lambda match: f"{match.group(1)}{int(match.group(2)):02d}{int(match.group(3)):02d}",
        text,
    )
    return re.sub(r"[\s,，。:：\-/.年月日]", "", text)


def _artifact(result: ChatResult, kind: str) -> dict[str, Any] | None:
    return next((item for item in result.artifacts if item.get("kind") == kind), None)


def _download_artifact(base_url: str, artifact: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    download_url = str(artifact["download_url"])
    url = download_url if download_url.startswith("http") else f"{base_url.rstrip('/')}{download_url}"
    suffix = {"word": ".docx", "excel": ".xlsx", "chart": ".svg"}.get(str(artifact.get("kind")), ".bin")
    output = output_dir / f"artifact-{artifact.get('id', 'unknown')}{suffix}"
    with request.urlopen(url, timeout=60) as response:
        output.write_bytes(response.read())
    return output


def grade(result: ChatResult, rule: dict[str, Any], base_url: str, output_dir: Path) -> Grade:
    grader_type = rule["type"]
    answer = _normalized(result.answer)
    values = [_normalized(value) for value in rule.get("values", [])]
    if grader_type == "contains_all":
        missing = [value for value in values if value not in answer]
        return Grade(not missing, "matched all required values" if not missing else f"missing values: {missing}")
    if grader_type == "contains_any":
        matched = [value for value in values if value in answer]
        return Grade(bool(matched), f"matched values: {matched}" if matched else f"none matched: {values}")

    artifact = _artifact(result, str(rule["kind"]))
    if not artifact:
        return Grade(False, f"missing {rule['kind']} artifact")
    path = _download_artifact(base_url, artifact, output_dir)
    if grader_type == "artifact_word":
        text = "\n".join(paragraph.text for paragraph in Document(path).paragraphs)
        missing = [value for value in values if value not in _normalized(text)]
        return Grade(not missing, "word artifact contains required values" if not missing else f"word missing: {missing}")
    if grader_type == "artifact_excel":
        workbook = load_workbook(path, data_only=True)
        cells = [_normalized(cell.value) for row in workbook.active.iter_rows() for cell in row if cell.value is not None]
        workbook.close()
        missing = [value for value in values if not any(value in cell for cell in cells)]
        return Grade(not missing, "excel artifact contains required values" if not missing else f"excel missing: {missing}")
    if grader_type == "artifact_excel_highlight":
        workbook = load_workbook(path)
        sheet = workbook.active
        rows = {
            _normalized(row[0].value): row[0].fill.fgColor.rgb == "FFFF9999"
            for row in sheet.iter_rows(min_row=2)
            if row and row[0].value is not None
        }
        workbook.close()
        expected = [_normalized(item) for item in rule["highlighted_first_column_values"]]
        unexpected = [_normalized(item) for item in rule["not_highlighted_first_column_values"]]
        passed = all(rows.get(item) is True for item in expected) and all(rows.get(item) is False for item in unexpected)
        return Grade(passed, f"highlight states: {rows}")
    if grader_type == "artifact_chart":
        text = _normalized(path.read_text(encoding="utf-8"))
        missing = [value for value in values if value not in text]
        return Grade(not missing, "chart artifact contains required values" if not missing else f"chart missing: {missing}")
    if grader_type == "artifact_spreadsheet_golden":
        return _grade_spreadsheet_golden(path, rule)
    return Grade(False, f"unknown grader type: {grader_type}")


def _grade_spreadsheet_golden(actual_path: Path, rule: dict[str, Any]) -> Grade:
    golden_path = Path(rule["golden_path"])
    position = str(rule["answer_position"])
    min_column, min_row, max_column, max_row = range_boundaries(position)
    actual_values = load_workbook(actual_path, data_only=True)
    actual_formulas = load_workbook(actual_path, data_only=False)
    golden_values = load_workbook(golden_path, data_only=True)
    golden_formulas = load_workbook(golden_path, data_only=False)
    sheet_name = str(rule.get("answer_sheet") or "").strip()
    actual_value_sheet = actual_values[sheet_name] if sheet_name else actual_values.active
    actual_formula_sheet = actual_formulas[sheet_name] if sheet_name else actual_formulas.active
    golden_value_sheet = golden_values[sheet_name] if sheet_name else golden_values.active
    golden_formula_sheet = golden_formulas[sheet_name] if sheet_name else golden_formulas.active
    mismatches = []
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            actual_value = actual_value_sheet.cell(row, column).value
            actual_formula = actual_formula_sheet.cell(row, column).value
            golden_value = golden_value_sheet.cell(row, column).value
            golden_formula = golden_formula_sheet.cell(row, column).value
            if not _excel_cell_matches(actual_value, actual_formula, golden_value, golden_formula):
                coordinate = actual_value_sheet.cell(row, column).coordinate
                if actual_value is None and isinstance(actual_formula, str) and actual_formula.startswith("="):
                    mismatches.append(f"{coordinate}: formula cache empty for {actual_formula!r}, expected={golden_value!r}")
                else:
                    mismatches.append(f"{coordinate}: actual={actual_value!r}, expected={golden_value!r}")
    actual_values.close()
    actual_formulas.close()
    golden_values.close()
    golden_formulas.close()
    if mismatches:
        return Grade(False, "golden mismatches: " + "; ".join(mismatches[:8]))
    return Grade(True, f"matched golden workbook range {position}")


def _excel_cell_matches(actual_value: Any, actual_formula: Any, golden_value: Any, golden_formula: Any) -> bool:
    if isinstance(golden_value, (date, datetime)) and isinstance(actual_value, str):
        return _normalized(actual_value) == golden_value.strftime("%Y%m%d")
    if isinstance(actual_value, (date, datetime)) and isinstance(golden_value, str):
        return actual_value.strftime("%Y%m%d") == _normalized(golden_value)
    if isinstance(actual_value, (int, float)) and isinstance(golden_value, (int, float)):
        return abs(float(actual_value) - float(golden_value)) < 1e-6
    if actual_value == golden_value:
        return True
    if actual_value is not None and golden_value is not None:
        return _normalized(actual_value) == _normalized(golden_value)
    return isinstance(actual_formula, str) and actual_formula == golden_formula
