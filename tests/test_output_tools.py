from pathlib import Path

from openpyxl import load_workbook

from app.tools import output_tools


def test_generate_word_and_excel(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)

    word = output_tools.generate_word("报告.docx", "第一段\n第二段")
    excel = output_tools.generate_excel("表格.xlsx", "产品,金额\nA,100")

    assert word.exists()
    assert word.suffix == ".docx"
    assert excel.exists()
    assert excel.suffix == ".xlsx"


def test_generate_excel_highlights_rows_above_threshold(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)

    excel = output_tools.generate_excel(
        "sales.xlsx",
        "product,amount\nA,100\nB,650\nC,500",
        highlight_gt=500,
        highlight_scope="row",
    )

    workbook = load_workbook(excel)
    sheet = workbook.active

    assert sheet["A2"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A3"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["B3"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["A4"].fill.fgColor.rgb != "FFFF9999"
    workbook.close()


def test_generate_excel_highlights_rows_below_threshold(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)

    excel = output_tools.generate_excel(
        "sales.xlsx",
        "product,amount\nA,900\nB,1200\nC,1000",
        highlight_lt=1000,
        highlight_scope="row",
    )

    workbook = load_workbook(excel)
    sheet = workbook.active

    assert sheet["A2"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["B2"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["A3"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["B3"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A4"].fill.fgColor.rgb != "FFFF9999"
    workbook.close()


def test_generate_excel_highlights_below_threshold_only_in_named_column(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)

    excel = output_tools.generate_excel(
        "sales.xlsx",
        "区域,销售额,增长率\n华东,1000,0.12\n华南,800,-0.05\n中部,1800,0.35\n华北,1500,0.22\n西南,1200,0.28\n东北,950,0.07",
        highlight_lt=1000,
        highlight_column="销售额",
        highlight_scope="row",
    )

    workbook = load_workbook(excel)
    sheet = workbook.active

    assert sheet["A2"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A3"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["B3"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["C3"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["A4"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A5"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A6"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A7"].fill.fgColor.rgb == "FFFF9999"
    workbook.close()


def test_generate_excel_can_highlight_only_cells(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)

    excel = output_tools.generate_excel(
        "sales.xlsx",
        "product,amount\nA,650",
        highlight_gt=500,
        highlight_scope="cell",
    )

    workbook = load_workbook(excel)
    sheet = workbook.active

    assert sheet["A2"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["B2"].fill.fgColor.rgb == "FFFF9999"
    workbook.close()
