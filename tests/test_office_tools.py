import json
from pathlib import Path
from email.message import EmailMessage

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.db.database import Base
from app.db.models import FileRecord, PdfTableRecord
from app.tools import email_tools
from app.tools import output_tools
from app.tools.office_tools import build_office_tools


def _tool_by_name(tools, name):
    return next(tool for tool in tools if tool.name == name)


def test_office_tools_read_file_and_generate_word(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        record = FileRecord(filename="demo.txt", path="demo.txt", extracted_text="项目背景：智能办公")
        db.add(record)
        db.commit()
        db.refresh(record)

        tools = build_office_tools(db, public_base_url="http://localhost:8000")
        read_result = json.loads(_tool_by_name(tools, "read_file").invoke({"file_id": record.id}))
        assert read_result["ok"] is True
        assert "智能办公" in read_result["file"]["content"]

        word_result = json.loads(
            _tool_by_name(tools, "generate_word_report").invoke({"title": "报告", "content": "正文"})
        )
        assert word_result["ok"] is True
        assert word_result["artifact"]["download_url"].startswith("/api/files/artifacts/")


def test_analyze_excel_tool(tmp_path: Path) -> None:
    xlsx = tmp_path / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售"
    sheet.append(["区域", "销售额"])
    sheet.append(["华东", 100])
    workbook.save(xlsx)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        record = FileRecord(filename="sales.xlsx", path=str(xlsx), extracted_text="区域 | 销售额")
        db.add(record)
        db.commit()
        db.refresh(record)

        result = json.loads(_tool_by_name(build_office_tools(db), "analyze_excel").invoke({"file_id": record.id}))

    assert result["ok"] is True
    assert result["analysis"]["sheets"][0]["sheet_name"] == "销售"
    assert result["analysis"]["sheets"][0]["headers"] == ["区域", "销售额"]


def test_generate_excel_chart_tool_returns_image_artifact(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)
    xlsx = tmp_path / "销售.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售"
    sheet.append(["月份", "收入", "成本"])
    sheet.append(["一月", 100, 70])
    sheet.append(["二月", 150, 90])
    sheet.append(["三月", 130, 80])
    workbook.save(xlsx)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        record = FileRecord(filename="销售.xlsx", path=str(xlsx), extracted_text="月份 | 收入 | 成本")
        db.add(record)
        db.commit()
        db.refresh(record)

        result = json.loads(
            _tool_by_name(build_office_tools(db), "generate_excel_chart").invoke(
                {
                    "file_id": record.id,
                    "chart_type": "line",
                    "x_axis_column": "月份",
                    "y_columns": ["收入"],
                    "title": "收入趋势",
                }
            )
        )

    artifact = result["artifact"]
    image_path = Path(artifact["path"])
    assert result["ok"] is True
    assert artifact["kind"] == "chart"
    assert artifact["download_url"].startswith("/api/files/artifacts/")
    assert artifact["metadata"]["chart_type"] == "line"
    assert artifact["metadata"]["x_axis_column"] == "月份"
    assert image_path.exists()
    assert image_path.suffix == ".svg"
    assert "收入趋势" in image_path.read_text(encoding="utf-8")


def test_pdf_table_tools_return_structured_table() -> None:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        record = FileRecord(filename="paper.pdf", path="paper.pdf", extracted_text="[page=5]\nTable 1")
        db.add(record)
        db.commit()
        db.refresh(record)
        db.add(
            PdfTableRecord(
                file_id=record.id,
                label="Table 1",
                caption="Table 1: Main results",
                page_number=5,
                data_json=json.dumps([["Model", "Score"], ["Ours", "0.91"]]),
                raw_text="Model | Score\nOurs | 0.91",
                extraction_method="test",
                confidence=1.0,
            )
        )
        db.commit()

        tools = build_office_tools(db)
        list_result = json.loads(_tool_by_name(tools, "list_pdf_tables").invoke({"file_id": record.id}))
        read_result = json.loads(
            _tool_by_name(tools, "read_pdf_table").invoke({"file_id": record.id, "table_label": "Table 1"})
        )
        missing_result = json.loads(
            _tool_by_name(tools, "read_pdf_table").invoke({"file_id": record.id, "table_label": "Table 2"})
        )

    assert list_result["ok"] is True
    assert list_result["tables"][0]["page"] == 5
    assert read_result["ok"] is True
    assert read_result["table"]["rows"][1] == ["Ours", "0.91"]
    assert missing_result["ok"] is False
    assert "available_tables" in missing_result


def test_generate_excel_tool_supports_row_highlight(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        tools = build_office_tools(db)
        result = json.loads(
            _tool_by_name(tools, "generate_excel_table").invoke(
                {
                    "filename": "sales.xlsx",
                    "content": "product,amount\nA,100\nB,650",
                    "highlight_gt": 500,
                    "highlight_scope": "row",
                }
            )
        )

    workbook = load_workbook(result["artifact"]["path"])
    sheet = workbook.active

    assert result["ok"] is True
    assert sheet["A2"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A3"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["B3"].fill.fgColor.rgb == "FFFF9999"
    workbook.close()


def test_generate_excel_tool_supports_less_than_row_highlight(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        tools = build_office_tools(db)
        result = json.loads(
            _tool_by_name(tools, "generate_excel_table").invoke(
                {
                    "filename": "sales.xlsx",
                    "content": "product,amount\nA,900\nB,1200",
                    "highlight_lt": 1000,
                    "highlight_scope": "row",
                }
            )
        )

    workbook = load_workbook(result["artifact"]["path"])
    sheet = workbook.active

    assert result["ok"] is True
    assert sheet["A2"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["B2"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["A3"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["B3"].fill.fgColor.rgb != "FFFF9999"
    workbook.close()


def test_generate_excel_tool_uses_named_highlight_column(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path)
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        tools = build_office_tools(db)
        result = json.loads(
            _tool_by_name(tools, "generate_excel_table").invoke(
                {
                    "filename": "sales.xlsx",
                    "content": "区域,销售额,增长率\n华东,1000,0.12\n华南,800,-0.05\n中部,1800,0.35\n东北,950,0.07",
                    "highlight_lt": 1000,
                    "highlight_column": "销售额",
                    "highlight_scope": "row",
                }
            )
        )

    workbook = load_workbook(result["artifact"]["path"])
    sheet = workbook.active

    assert result["ok"] is True
    assert sheet["A2"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A3"].fill.fgColor.rgb == "FFFF9999"
    assert sheet["A4"].fill.fgColor.rgb != "FFFF9999"
    assert sheet["A5"].fill.fgColor.rgb == "FFFF9999"
    workbook.close()


def test_edit_excel_cells_tool_copies_workbook_and_updates_cells(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path / "artifacts")
    monkeypatch.setattr(output_tools, "recalculate_excel_file", lambda path: True)
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet["A1"] = "名称"
    sheet["B2"] = "旧值"
    workbook.save(source)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        record = FileRecord(filename="source.xlsx", path=str(source), extracted_text="名称")
        db.add(record)
        db.commit()
        db.refresh(record)
        result = json.loads(
            _tool_by_name(build_office_tools(db), "edit_uploaded_excel_cells").invoke(
                {
                    "file_id": record.id,
                    "filename": "edited.xlsx",
                    "updates": [{"sheet_name": "数据", "cell": "B2", "value": "新值"}],
                }
            )
        )

    edited = load_workbook(result["artifact"]["path"])
    assert result["ok"] is True
    assert result["artifact"]["kind"] == "excel"
    assert edited["数据"]["B2"].value == "新值"
    edited.close()


def test_excel_range_tools_read_calculate_lookup_filter_and_write(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path / "artifacts")
    monkeypatch.setattr(output_tools, "recalculate_excel_file", lambda path: True)
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["team", "amount"])
    sheet.append(["A", 10])
    sheet.append(["B", 20])
    sheet.append(["A", 30])
    workbook.save(source)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        record = FileRecord(filename="source.xlsx", path=str(source), extracted_text="team | amount")
        db.add(record)
        db.commit()
        db.refresh(record)
        tools = build_office_tools(db)
        read_result = json.loads(
            _tool_by_name(tools, "read_excel_range").invoke({"file_id": record.id, "cell_range": "A1:B4"})
        )
        sum_result = json.loads(
            _tool_by_name(tools, "calculate_excel_sum").invoke(
                {"file_id": record.id, "sum_range": "B2:B4", "criteria_range": "A2:A4", "criteria": "A"}
            )
        )
        lookup_result = json.loads(
            _tool_by_name(tools, "lookup_excel").invoke(
                {"file_id": record.id, "lookup_value": "B", "lookup_range": "A2:A4", "result_range": "B2:B4"}
            )
        )
        filter_result = json.loads(
            _tool_by_name(tools, "filter_excel").invoke(
                {"file_id": record.id, "data_range": "A1:B4", "column": "amount", "operator": "gte", "value": 20}
            )
        )
        write_result = json.loads(
            _tool_by_name(tools, "write_uploaded_excel_range").invoke(
                {"file_id": record.id, "filename": "written.xlsx", "start_cell": "D2", "values": [[40], [50]]}
            )
        )

    assert read_result["rows"][2] == ["B", 20]
    assert sum_result["result"] == 40
    assert lookup_result["result"] == 20
    assert filter_result["matched_count"] == 2
    written = load_workbook(write_result["artifact"]["path"])
    assert written.active["D2"].value == 40
    assert written.active["D3"].value == 50
    written.close()


def test_fill_excel_formula_tool_translates_formula(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(output_tools.settings, "artifact_dir", tmp_path / "artifacts")
    monkeypatch.setattr(output_tools, "recalculate_excel_file", lambda path: True)
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A2"] = 3
    sheet["A3"] = 4
    workbook.save(source)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        record = FileRecord(filename="source.xlsx", path=str(source), extracted_text="values")
        db.add(record)
        db.commit()
        db.refresh(record)
        result = json.loads(
            _tool_by_name(build_office_tools(db), "fill_uploaded_excel_formula").invoke(
                {"file_id": record.id, "filename": "formula.xlsx", "cell_range": "B2:B3", "formula": "=A2*2"}
            )
        )

    edited = load_workbook(result["artifact"]["path"], data_only=False)
    assert edited.active["B2"].value == "=A2*2"
    assert edited.active["B3"].value == "=A3*2"
    edited.close()


def test_send_email_tool_uses_configured_smtp_with_attachments(tmp_path: Path, monkeypatch) -> None:
    sent_messages = []
    smtp_logins = []

    class FakeSMTP:
        def __init__(self, host, port, timeout=None):
            self.host = host
            self.port = port
            self.timeout = timeout

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def login(self, username, password):
            smtp_logins.append((username, password))

        def send_message(self, message):
            sent_messages.append(message)

    monkeypatch.setattr(email_tools.settings, "email_smtp_host", "smtp.qq.com")
    monkeypatch.setattr(email_tools.settings, "email_smtp_port", 465)
    monkeypatch.setattr(email_tools.settings, "email_smtp_username", "sender@qq.com")
    monkeypatch.setattr(email_tools.settings, "email_smtp_password", "authorization-code")
    monkeypatch.setattr(email_tools.settings, "email_from", "sender@qq.com")
    monkeypatch.setattr(email_tools.settings, "email_use_ssl", True)
    monkeypatch.setattr(email_tools.smtplib, "SMTP_SSL", FakeSMTP)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    attachment_path = tmp_path / "note.txt"
    attachment_path.write_text("attached content", encoding="utf-8")

    with Session(engine) as db:
        record = FileRecord(
            filename="note.txt",
            path=str(attachment_path),
            content_type="text/plain",
            extracted_text="attached content",
        )
        db.add(record)
        db.commit()
        db.refresh(record)

        result = json.loads(
            _tool_by_name(build_office_tools(db), "send_email").invoke(
                {
                    "to": "1254543711@qq.com",
                    "subject": "提醒",
                    "content": "明天来找我。",
                    "file_ids": [record.id],
                }
            )
        )

    assert result["ok"] is True
    assert result["email"]["status"] == "sent"
    assert result["email"]["to"] == "1254543711@qq.com"
    assert result["email"]["attachments"] == [{"filename": "note.txt", "content_type": "text/plain"}]
    assert smtp_logins == [("sender@qq.com", "authorization-code")]
    assert sent_messages[0]["From"] == "sender@qq.com"
    assert sent_messages[0]["To"] == "1254543711@qq.com"
    assert sent_messages[0]["Subject"] == "提醒"
    assert "明天来找我。" in sent_messages[0].get_body(preferencelist=("plain",)).get_content()
    attachments = list(sent_messages[0].iter_attachments())
    assert attachments[0].get_filename() == "note.txt"
    assert attachments[0].get_content() == "attached content"


def test_fetch_unread_emails_tool_returns_body_and_text_attachments(monkeypatch) -> None:
    message = EmailMessage()
    message["From"] = "sender@example.com"
    message["To"] = "receiver@qq.com"
    message["Subject"] = "Project update"
    message["Date"] = "Thu, 28 May 2026 10:00:00 +0800"
    message.set_content("Please review the attached plan.")
    message.add_attachment("Plan line 1\nPlan line 2", subtype="plain", filename="plan.txt")
    raw_message = message.as_bytes()
    fetch_modes = []
    store_calls = []

    class FakeIMAP:
        def __init__(self, host, port):
            self.host = host
            self.port = port

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def login(self, username, password):
            self.username = username
            self.password = password
            return "OK", [b"logged in"]

        def select(self, mailbox):
            return "OK", [b"1"]

        def search(self, charset, criteria):
            return "OK", [b"42"]

        def fetch(self, email_id, mode):
            fetch_modes.append(mode)
            return "OK", [(b"42 (BODY[])", raw_message)]

        def store(self, email_id, command, flags):
            store_calls.append((email_id, command, flags))
            return "OK", [b"marked seen"]

        def logout(self):
            return "OK", [b"logged out"]

    monkeypatch.setattr(email_tools.settings, "email_imap_host", "imap.qq.com")
    monkeypatch.setattr(email_tools.settings, "email_imap_port", 993)
    monkeypatch.setattr(email_tools.settings, "email_imap_username", "receiver@qq.com")
    monkeypatch.setattr(email_tools.settings, "email_imap_password", "authorization-code")
    monkeypatch.setattr(email_tools.settings, "email_mark_read_on_fetch", True)
    monkeypatch.setattr(email_tools.imaplib, "IMAP4_SSL", FakeIMAP)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as db:
        result = json.loads(_tool_by_name(build_office_tools(db), "fetch_unread_emails").invoke({}))

    email = result["unread_emails"]["emails"][0]
    assert result["ok"] is True
    assert result["unread_emails"]["count"] == 1
    assert result["unread_emails"]["marked_read"] is True
    assert email["from"] == "sender@example.com"
    assert email["subject"] == "Project update"
    assert "Please review the attached plan." in email["body"]
    assert email["attachments"][0]["filename"] == "plan.txt"
    assert "Plan line 1" in email["attachments"][0]["content_preview"]
    assert fetch_modes == ["(BODY.PEEK[])"]
    assert store_calls == [(b"42", "+FLAGS", "\\Seen")]
